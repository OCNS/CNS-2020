from openpyxl import load_workbook
from openpyxl.styles import Alignment
import html2text
import xml.etree.ElementTree as ET
import re


def format_xml_abstract(xml_text):
    parsed_xml = ET.fromstring('<root>' + xml_text + '</root>') # add a dummy root node before processing
    abstract = ''
    for child in parsed_xml.iter():
        if child.tag == 'root':
            if (child.text is not None) and (child.text.strip() != ''):
                raise Exception('%s node has text which is being ignored' % child.tag)
            else:
                continue
        elif child.tag == 'p':
            if (child.text is not None) and (child.text.strip() != ''):
                raise Exception('%s node has text which is being ignored' % child.tag)
            else:
                continue
        elif child.tag == 'span':
            if (child.text is not None) and (child.text.strip() != ''):
                print(child.text.replace('\n', ' ').replace('\r', ' '))

        # for child in p_child.iter():
        #     print(child.text)

    pass


def format_html_abstract(h, html_text, p):
    abstract = h.handle(html_text)
    abstract = p.sub(' ', abstract)
    abstract = abstract.replace('\n\n', '</br></br>')

    return abstract.strip()


def format_html_abstract_with_authors(h, html_text, paper_authors, p_abstract):
    abstract = h.handle(html_text)
    abstract = abstract.strip() # remove any white space before and after the abstract
    abstract = p_abstract.sub('</br></br>', abstract) # replace new paragraphs by two line breaks
    abstract = re.sub('\*\*([a-zA-Z0-9]+)\*\*', r'<strong>\1</strong>', abstract) # replace anything betweeb ** ... ** with corresponding text in bold
    # abstract = abstract.replace('\n\n', '</br></br>') # replace new paragraphs by two line breaks
    abstract = abstract.replace('\n', ' ') # replace any other line breaks by space

    authors = paper_authors.split(';')
    author_string = ''
    for i, author in enumerate(authors):
        author_string += '<strong>' + author + '</strong>'
        if i < (len(authors) - 1):
            author_string += ', '

    abstract = author_string + '</br></br>' + abstract # add a line after author's names

    return abstract


def format_author(author_text, p):
    author_text = author_text.replace(',', ';')
    author_text = p.sub('', author_text)

    return author_text


def begin_format1():
    sched_wb_path = 'data/restore_file.xlsx' # path to excel workbook for import in Schedback
    paper_wb_path = 'data/CNS_2020_Paper-2020-06-11.xlsx' # path to CNS papers workbook
    user_wb_path = 'data/CNS_2020_User-2020-06-11.xlsx' # path to CNS users workbook

    # load excel workbooks for processing
    sched_wb = load_workbook(filename=sched_wb_path)
    paper_wb = load_workbook(filename=paper_wb_path)
    user_wb = load_workbook(filename=user_wb_path)

    # setup html2text options
    h = html2text.HTML2Text()

    # setup parameters below carefully to ensure things are inserted correctly in Sched
    sched_first_row = 13 # begin inserting sessions in this row

    # regular expressions to format author names
    p = re.compile('\s\(#[0-9]*\)') # for formatting author names
    p_abstract = re.compile('\\n') # for formatting abstract

    last_session_id = 4 # ID of the last session in the excel sheet
    paper_ws = paper_wb['Tablib Dataset'] # access paper worksheet
    sched_ws = sched_wb['Sessions'] # access sessions worksheet
    poster_count = 0 # keep track of count of posters to assign them to separate time slots
    poster_start_day = 19 # start day for poster sessions
    poster_start_time = 7 # start time for poster sessions
    for row in paper_ws.iter_rows(min_row=2): # ignore the first row that contains column names
        paper_type = row[3].value

        if paper_type != 'Rejected':
            # extract information about the paper
            paper_ID = str(last_session_id + 1)
            paper_abstract = format_html_abstract(h, row[10].value, p_abstract)
            paper_title = row[9].value
            paper_author = format_author(row[1].value, p)

            # add paper details to sched workbook
            sched_ws['A%d' % sched_first_row] = paper_ID
            sched_ws['B%d' % sched_first_row] = paper_title
            sched_ws['C%d' % sched_first_row] = 'Y'
            sched_ws['C%d' % sched_first_row].alignment = Alignment(horizontal='center')
            sched_ws['D%d' % sched_first_row] = 'N'
            sched_ws['J%d' % sched_first_row] = paper_abstract
            sched_ws['K%d' % sched_first_row] = paper_author
            sched_ws['P%d' % sched_first_row] = 'TBA'

            if paper_type == 'Accepted':
                sched_ws['G%d' % sched_first_row] = 'Poster'

                # start and end times for posters
                poster_count = poster_count + 1

                if poster_count % 120 == 0:
                    poster_start_day = poster_start_day + 1
                    poster_start_time = 7
                elif poster_count % 40 == 0:
                    poster_start_time = poster_start_time + 1

                sched_ws['E%d' % sched_first_row] = '7/%d/2020 %d:00 PM' % (poster_start_day, poster_start_time) # start time for posters
                sched_ws['F%d' % sched_first_row] = '7/%d/2020 %d:00 PM' % (poster_start_day, poster_start_time + 1) # end time for posters
                sched_ws['F%d' % sched_first_row].alignment = Alignment(horizontal='right')
            elif paper_type == 'AcceptedOral':
                sched_ws['G%d' % sched_first_row] = 'Oral'
            elif paper_type == 'AcceptedFeatured':
                sched_ws['G%d' % sched_first_row] = 'Featured Talk'

            sched_first_row = sched_first_row + 1
            last_session_id = last_session_id + 1

    sched_wb.save(sched_wb_path)


def get_oral_sessions_time(session):
    oral_sessions = dict()
    oral_sessions['Using evolutionary algorithms to explore single-cell heterogeneity and microcircuit operation in the hippocampus'] = ('7/19/2020 01:00 PM', '7/19/2020 01:40 PM')
    oral_sessions['\'Awake Delta\' and Theta-Rhythmic Modes of Hippocampal Network Activity Track Intermittent Locomotor Behaviors in Rat'] = ('7/19/2020 01:40 PM', '7/19/2020 02:00 PM')
    oral_sessions['Neural Manifold Models for Characterising Brain Circuit Dynamics in Neurodegenerative Disease'] = ('7/19/2020 02:00 PM', '7/19/2020 02:20 PM')
    oral_sessions['Coupled experimental and modeling representation of the mechanisms of epileptic discharges in rat brain slices'] = ('7/19/2020 02:20 PM', '7/19/2020 02:40 PM')

    oral_sessions['Towards multipurpose bio-realistic models of cortical circuits'] = ('7/19/2020 04:20 PM', '7/19/2020 04:40 PM')
    oral_sessions['How Stimulus Statistics Affect the Receptive Fields of Cells in Primary Visual Cortex'] = ('7/19/2020 04:40 PM', '7/19/2020 05:00 PM')
    oral_sessions['Analysis and Modelling of Response Features of Accessory Olfactory Bulb Neurons'] = ('7/19/2020 05:00 PM', '7/19/2020 05:20 PM')

    oral_sessions['Delineating Reward/Avoidance Decision Process in the Impulsive-compulsive Spectrum Disorders through a Probabilistic Reversal Learning Task'] = ('7/19/2020 05:40 PM', '7/19/2020 06:20 PM')
    oral_sessions['Dopamine role in learning and action inference'] = ('7/19/2020 06:20 PM', '7/19/2020 06:40 PM')

    oral_sessions['Neuronal morphology imposes a tradeoff between stability, accuracy and efficiency of synaptic scaling'] = ('7/20/2020 01:00 PM', '7/20/2020 01:40 PM')
    oral_sessions['Finite element simulation of ionic electrodiffusion in cellular geometries'] = ('7/20/2020 01:40 PM', '7/20/2020 02:00 PM')
    oral_sessions['Discovering synaptic mechanisms underlying the propagation of cortical activity: A model-driven experimental and data analysis approach'] = ('7/20/2020 02:00 PM', '7/20/2020 02:20 PM')
    oral_sessions['Neural flows: estimation of wave velocities and identification of singularities in 3D+t brain data'] = ('7/20/2020 02:20 PM', '7/20/2020 02:40 PM')

    oral_sessions['Who can turn faster? Comparison of the head direction circuit of two species'] = ('7/20/2020 04:20 PM', '7/20/2020 05:00 PM')
    oral_sessions['Experimental and computational characterization of interval variability in the sequential activity of the Lymnaea feeding CPG'] = ('7/20/2020 05:00 PM', '7/20/2020 05:20 PM')

    oral_sessions['A Spatial Developmental Generative Model of Human Brain Structural Connectivity'] = ('7/20/2020 05:40 PM', '7/20/2020 06:00 PM')
    oral_sessions['Cortical integration and segregation explained by harmonic modes of functional connectivity'] = ('7/20/2020 06:00 PM', '7/20/2020 06:20 PM')
    oral_sessions['Reconciling emergences: An information-theoretic approach to identify causal emergence in multivariate data'] = ('7/20/2020 06:20 PM', '7/20/2020 06:40 PM')

    return oral_sessions[session]


def begin_format2():
    sched_wb_path = 'data/restore_file.xlsx' # path to excel workbook for import in Schedback
    paper_wb_path = 'data/CNS_2020_Paper-2020-06-11.xlsx' # path to CNS papers workbook
    user_wb_path = 'data/CNS_2020_User-2020-06-11.xlsx' # path to CNS users workbook

    # load excel workbooks for processing
    sched_wb = load_workbook(filename=sched_wb_path)
    paper_wb = load_workbook(filename=paper_wb_path)
    user_wb = load_workbook(filename=user_wb_path)

    # setup html2text options
    h = html2text.HTML2Text()

    # setup parameters below carefully to ensure things are inserted correctly in Sched
    sched_first_row = 13 # begin inserting sessions in this row

    # regular expressions to format author names
    p_authors = re.compile('\s\(#[0-9]*\)') # for formatting author names
    p_abstract = re.compile('\n\n([\n\s])*') # for formatting abstract

    last_session_id = 4  # ID of the last session in the excel sheet
    paper_ws = paper_wb['Tablib Dataset'] # access paper worksheet
    sched_ws = sched_wb['Sessions'] # access sessions worksheet
    poster_count = 0 # keep track of count of posters to assign them to separate time slots
    poster_start_day = 19 # start day for poster sessions
    poster_start_time = 7 # start time for poster sessions
    for row in paper_ws.iter_rows(min_row=2): # ignore the first row that contains column names
        paper_type = row[3].value

        if paper_type != 'Rejected':
            # extract information about the paper
            paper_ID = str(last_session_id + 1)
            paper_title = row[9].value
            paper_author = format_author(row[1].value, p_authors)
            paper_speaker = format_author(row[2].value, p_authors)
            paper_abstract = format_html_abstract_with_authors(h, row[10].value, paper_author, p_abstract)

            # add paper details to sched workbook
            sched_ws['A%d' % sched_first_row] = paper_ID
            sched_ws['B%d' % sched_first_row] = paper_title
            sched_ws['C%d' % sched_first_row] = 'Y'
            sched_ws['C%d' % sched_first_row].alignment = Alignment(horizontal='center')
            sched_ws['D%d' % sched_first_row] = 'N'
            sched_ws['J%d' % sched_first_row] = paper_abstract
            sched_ws['K%d' % sched_first_row] = paper_speaker
            sched_ws['P%d' % sched_first_row] = 'TBA'

            if paper_type == 'Accepted':
                sched_ws['G%d' % sched_first_row] = 'Poster'

                # start and end times for posters
                poster_count = poster_count + 1

                if poster_count % 120 == 0:
                    poster_start_day = poster_start_day + 1
                    poster_start_time = 7
                elif poster_count % 40 == 0:
                    poster_start_time = poster_start_time + 1

                sched_ws['E%d' % sched_first_row] = '7/%d/2020 %d:00 PM' % (poster_start_day, poster_start_time) # start time for posters
                sched_ws['F%d' % sched_first_row] = '7/%d/2020 %d:00 PM' % (poster_start_day, poster_start_time + 1) # end time for posters
                sched_ws['F%d' % sched_first_row].alignment = Alignment(horizontal='right')
            elif paper_type == 'AcceptedOral':
                sched_ws['G%d' % sched_first_row] = 'Oral'
                start_time, end_time = get_oral_sessions_time(paper_title)
                sched_ws['E%d' % sched_first_row] = start_time  # start time for oral session
                sched_ws['F%d' % sched_first_row] = end_time  # end time for oral session
                sched_ws['F%d' % sched_first_row].alignment = Alignment(horizontal='right')
            elif paper_type == 'AcceptedFeatured':
                sched_ws['G%d' % sched_first_row] = 'Featured Talk'
                start_time, end_time = get_oral_sessions_time(paper_title)
                sched_ws['E%d' % sched_first_row] = start_time  # start time for oral session
                sched_ws['F%d' % sched_first_row] = end_time  # end time for oral session
                sched_ws['F%d' % sched_first_row].alignment = Alignment(horizontal='right')

            sched_first_row = sched_first_row + 1
            last_session_id = last_session_id + 1

    sched_wb.save(sched_wb_path)


def update_speaker_profile():
    sched_wb_path = 'data/cns2020online-speakers-directory-2020-06-20-00-53-57.xlsx'
    user_wb_path = 'data/CNS_2020_User-2020-06-11.xlsx'  # path to CNS users workbook

    # load excel workbooks for processing
    sched_wb = load_workbook(filename=sched_wb_path)
    user_wb = load_workbook(filename=user_wb_path)

    # load specific worksheets for processing
    user_ws = user_wb['Tablib Dataset']  # access paper worksheet
    sched_ws = sched_wb['Speakers']  # access sessions worksheet

    for row in user_ws.iter_rows(min_row=2):
        author_name = row[1]. value + ' ' + row[2].value

        # search for author in Sched worksheet
        min_row = 6
        author_found = False
        for i, sched_row in enumerate(sched_ws.iter_rows(min_row=min_row)):
            if author_name == sched_row[0].value:
                author_found = True
                break

        if author_found:
            # add author email
            sched_ws['B%d' % (i + min_row)] = row[9].value

            # add author company
            if sched_row[3].value is None:
                if row[4].value and row[3].value:
                    sched_ws['D%d' % (i + min_row)] = row[4].value + ', ' + row[3].value
                elif row[3].value:
                    sched_ws['D%d' % (i + min_row)] = row[3].value

    # save workbook
    sched_wb.save(sched_wb_path)


if __name__ == '__main__':
    # begin_format2()
    update_speaker_profile()